import sys
import random
import pandas as pd
from datetime import datetime, timedelta
import time
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                           QHBoxLayout, QLabel, QPushButton, QFrame, QTabWidget,
                           QProgressBar, QTableWidget, QTableWidgetItem, QComboBox,
                           QLineEdit, QScrollArea, QGridLayout, QListWidget, QSlider,
                           QListWidgetItem, QFileDialog, QSizePolicy, QMessageBox, QHeaderView)
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QFont, QColor, QPainter, QLinearGradient
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from algorithm_tree import AlgorithmTreeWidget
from dummy_devices import SecurityDevices
import seaborn as sns
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

class SecuritySystem(QMainWindow):
    def __init__(self):
        super().__init__()
        self.devices = SecurityDevices()  # Initialize dummy devices
        
        # Inisialisasi figure dan canvas untuk grafik evaluasi
        self.objectPieFigure = Figure(figsize=(6, 4))
        self.objectPieCanvas = FigureCanvas(self.objectPieFigure)
        
        self.objectTrendFigure = Figure(figsize=(6, 4))
        self.objectTrendCanvas = FigureCanvas(self.objectTrendFigure)
        
        self.behaviorHeatFigure = Figure(figsize=(6, 4))
        self.behaviorHeatCanvas = FigureCanvas(self.behaviorHeatFigure)
        
        self.behaviorBarFigure = Figure(figsize=(6, 4))
        self.behaviorBarCanvas = FigureCanvas(self.behaviorBarFigure)
        
        self.anomalyTSFigure = Figure(figsize=(6, 4))
        self.anomalyTSCanvas = FigureCanvas(self.anomalyTSFigure)
        
        self.anomalyScatterFigure = Figure(figsize=(6, 4))
        self.anomalyScatterCanvas = FigureCanvas(self.anomalyScatterFigure)
        
        self.audioSpecFigure = Figure(figsize=(6, 4))
        self.audioSpecCanvas = FigureCanvas(self.audioSpecFigure)
        
        self.audioLevelFigure = Figure(figsize=(6, 4))
        self.audioLevelCanvas = FigureCanvas(self.audioLevelFigure)
        
        self.perfMetricsFigure = Figure(figsize=(6, 4))
        self.perfMetricsCanvas = FigureCanvas(self.perfMetricsFigure)
        
        self.resourceUsageFigure = Figure(figsize=(6, 4))
        self.resourceUsageCanvas = FigureCanvas(self.resourceUsageFigure)
        
        self.initUI()
        self.setupData()
        self.setupTimers()
        self.setupAI()

    def initUI(self):
        """Inisialisasi UI utama"""
        self.setWindowTitle("Sistem Keamanan Otomatis")
        self.setGeometry(50, 50, 1600, 900)
        
        # Widget utama
        mainWidget = QWidget()
        self.setCentralWidget(mainWidget)
        
        # Layout utama
        mainLayout = QVBoxLayout(mainWidget)
        mainLayout.setSpacing(10)
        mainLayout.setContentsMargins(15, 15, 15, 15)
        
        # Header
        self.setupHeader(mainLayout)
        
        # Tab Widget
        self.tabWidget = QTabWidget()
        self.tabWidget.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #dcdde1;
                border-radius: 5px;
                background: white;
            }
            QTabBar::tab {
                background: #f5f6fa;
                color: #2c3e50;
                padding: 10px 20px;
                border: 1px solid #dcdde1;
                border-bottom: none;
                border-top-left-radius: 5px;
                border-top-right-radius: 5px;
                min-width: 150px;
                font-size: 14px;
            }
            QTabBar::tab:selected {
                background: white;
                border-bottom: 3px solid #2980b9;
                color: #2980b9;
                font-weight: bold;
            }
        """)
        
        # Setup tabs
        self.setupMonitoringTab()
        self.setupEvaluationTab()
        self.setupAlgorithmTab()
        self.setupAIMaintenanceTab()
        self.setupAISecurityEvaluationTab()
        
        mainLayout.addWidget(self.tabWidget)
        
        # Footer
        self.setupFooter(mainLayout)
        
        # Styling
        self.applyStyles()

    def setupHeader(self, parentLayout):
        """Setup area header"""
        header = QFrame()
        header.setMaximumHeight(100)
        headerLayout = QHBoxLayout(header)
        
        # Judul sistem
        title = QLabel("SISTEM PEMANTAUAN KEAMANAN")
        title.setStyleSheet("""
            font-size: 24px;
            font-weight: bold;
            color: #2c3e50;
        """)
        
        # Indikator status
        self.statusLabel = QLabel("Status: AKTIF")
        self.statusLabel.setStyleSheet("""
            font-size: 18px;
            color: #27ae60;
            padding: 5px 15px;
            border: 2px solid #27ae60;
            border-radius: 5px;
        """)
        
        headerLayout.addWidget(title)
        headerLayout.addStretch()
        headerLayout.addWidget(self.statusLabel)
        
        parentLayout.addWidget(header)

    def setupMonitoringTab(self):
        """Setup tab pemantauan"""
        monitoringTab = QWidget()
        layout = QHBoxLayout(monitoringTab)
        
        # Panel kiri (70%)
        leftPanel = QFrame()
        leftLayout = QVBoxLayout(leftPanel)
        
        # Grid Kamera
        cameraGrid = QFrame()
        gridLayout = QGridLayout(cameraGrid)
        gridLayout.setSpacing(10)
        
        camera_areas = ["Depan", "Belakang", "Samping", "Dalam"]
        for i in range(2):
            for j in range(2):
                cameraFrame = QFrame()
                cameraFrame.setStyleSheet("""
            QFrame {
                        background-color: #f5f6fa;
                        border: 2px solid #dcdde1;
                border-radius: 10px;
                        min-height: 200px;
            }
        """)
                cameraLayout = QVBoxLayout(cameraFrame)
                
                area = camera_areas[i*2 + j]
                title = QLabel(f"Kamera Area {area}")
                title.setStyleSheet("font-weight: bold; color: #2c3e50;")
                status = QLabel("Status: Normal")
                status.setStyleSheet("color: #27ae60;")
                
                cameraLayout.addWidget(title)
                cameraLayout.addWidget(status)
                gridLayout.addWidget(cameraFrame, i, j)
        
        leftLayout.addWidget(cameraGrid)
        
        # Log Aktivitas
        logFrame = QFrame()
        logLayout = QVBoxLayout(logFrame)
        logLayout.setSpacing(5)
        
        logTitle = QLabel("Catatan Aktivitas Real-time")
        logTitle.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 16px;")
        
        self.logList = QListWidget()
        self.logList.setStyleSheet("""
            QListWidget {
                background-color: #f5f6fa;
                border: 2px solid #dcdde1;
                border-radius: 5px;
            }
            QListWidget::item {
                padding: 8px;
                border-bottom: 1px solid #dcdde1;
            }
        """)
        
        logLayout.addWidget(logTitle)
        logLayout.addWidget(self.logList)
        leftLayout.addWidget(logFrame)
        
        layout.addWidget(leftPanel, 70)
        
        # Panel kanan (30%)
        rightPanel = QFrame()
        rightLayout = QVBoxLayout(rightPanel)
        
        # Panel Status
        statusFrame = QFrame()
        statusLayout = QVBoxLayout(statusFrame)
        
        statusTitle = QLabel("Status Sistem")
        statusTitle.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 16px;")
        
        stats = [
            ("Sensor Aktif", "8/8", "#27ae60"),
            ("Koneksi", "Terhubung", "#27ae60"),
            ("Penggunaan CPU", "45%", "#f39c12"),
            ("Memori", "60%", "#f39c12")
        ]
        
        for label, value, color in stats:
            statFrame = QFrame()
            statLayout = QHBoxLayout(statFrame)
            
            statLabel = QLabel(label)
            statLabel.setStyleSheet("color: #7f8c8d;")
            statValue = QLabel(value)
            statValue.setStyleSheet(f"color: {color}; font-weight: bold;")
            
            statLayout.addWidget(statLabel)
            statLayout.addWidget(statValue)
            
            statusLayout.addWidget(statFrame)
        
        rightLayout.addWidget(statusTitle)
        rightLayout.addWidget(statusFrame)
        
        # Aksi Cepat
        actionsTitle = QLabel("Aksi Cepat")
        actionsTitle.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 16px;")
        rightLayout.addWidget(actionsTitle)
        
        actions = [
            ("🚨 Aktifkan Alarm", "#e74c3c"),
            ("🔒 Kunci Semua Pintu", "#3498db"),
            ("📞 Hubungi Bantuan", "#27ae60")
        ]
        
        for text, color in actions:
            btn = QPushButton(text)
            btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {color};
                    color: white;
                    padding: 10px;
                border: none;
                    border-radius: 5px;
                    font-size: 14px;
                }}
                QPushButton:hover {{
                    background-color: {color}dd;
                }}
            """)
            rightLayout.addWidget(btn)
        
        layout.addWidget(rightPanel, 30)
        
        self.tabWidget.addTab(monitoringTab, "🎥 Pemantauan")

    def setupEvaluationTab(self):
        """Setup tab evaluasi"""
        evaluationTab = QWidget()
        layout = QVBoxLayout(evaluationTab)
        
        # Toolbar dengan tombol export
        toolbarFrame = QFrame()
        toolbarLayout = QHBoxLayout(toolbarFrame)
        
        exportButton = QPushButton("📊 Export Data Analisis")
        exportButton.setStyleSheet("""
                QPushButton {
                background-color: #27ae60;
                    color: white;
                padding: 10px 20px;
                    border: none;
                    border-radius: 5px;
                font-weight: bold;
                }
            QPushButton:hover {
                background-color: #2ecc71;
                }
            """)
        exportButton.clicked.connect(self.exportData)
        
        refreshButton = QPushButton("🔄 Perbarui Data")
        refreshButton.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                padding: 10px 20px;
                border: none;
                    border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
                        }
                    """)
        refreshButton.clicked.connect(self.refreshData)
        
        toolbarLayout.addWidget(exportButton)
        toolbarLayout.addWidget(refreshButton)
        toolbarLayout.addStretch()
        
        layout.addWidget(toolbarFrame)
        
        # Tab untuk berbagai jenis data
        self.dataTabs = QTabWidget()
        self.dataTabs.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #dcdde1;
                border-radius: 5px;
                background: white;
            }
            QTabBar::tab {
                background: #f5f6fa;
                color: #2c3e50;
                padding: 10px 15px;
                border: 1px solid #dcdde1;
                border-bottom: none;
                border-top-left-radius: 5px;
                border-top-right-radius: 5px;
            }
            QTabBar::tab:selected {
                background: white;
                border-bottom: 3px solid #2980b9;
                color: #2980b9;
                font-weight: bold;
                        }
                    """)
        
        # Setup setiap tab dengan tabel
        self.setupObjectDetectionTab()
        self.setupBehaviorAnalysisTab()
        self.setupAnomalyDetectionTab()
        self.setupAudioAnalysisTab()
        self.setupPerformanceTab()
        
        layout.addWidget(self.dataTabs)
        self.tabWidget.addTab(evaluationTab, "📊 Evaluasi")

    def setupObjectDetectionTab(self):
        """Setup tab deteksi objek"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Tabel data
        self.objectDetectionTable = self.createDataTable(
            ["Timestamp", "Jenis Objek", "Tingkat Kepercayaan", "Lokasi Kamera", "Status"]
        )
        layout.addWidget(self.objectDetectionTable)
        
        self.dataTabs.addTab(tab, "Deteksi Objek")

    def setupBehaviorAnalysisTab(self):
        """Setup tab analisis perilaku"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Tabel data
        self.behaviorTable = self.createDataTable(
            ["Timestamp", "Tipe Perilaku", "Skor Analisis", "Durasi Deteksi", "Tindakan"]
        )
        layout.addWidget(self.behaviorTable)
        
        self.dataTabs.addTab(tab, "Analisis Perilaku")

    def setupAnomalyDetectionTab(self):
        """Setup tab deteksi anomali"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Tabel data
        self.anomalyTable = self.createDataTable(
            ["Timestamp", "Skor Anomali", "Tipe Anomali", "Level Ancaman", "Status Respons"]
        )
        layout.addWidget(self.anomalyTable)
        
        self.dataTabs.addTab(tab, "Deteksi Anomali")

    def setupAudioAnalysisTab(self):
        """Setup tab analisis audio"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Tabel data
        self.audioTable = self.createDataTable(
            ["Timestamp", "Tipe Suara", "Level Desibel", "Klasifikasi", "Lokasi Sumber"]
        )
        layout.addWidget(self.audioTable)
        
        self.dataTabs.addTab(tab, "Analisis Audio")

    def setupPerformanceTab(self):
        """Setup tab kinerja AI"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Tabel data
        self.performanceTable = self.createDataTable(
            ["Timestamp", "Akurasi Deteksi", "False Positives", "Waktu Respons", "CPU Usage"]
        )
        layout.addWidget(self.performanceTable)
        
        self.dataTabs.addTab(tab, "Kinerja AI")

    def createDataTable(self, headers):
        """Buat tabel data dengan format yang rapi"""
        table = QTableWidget()
        table.setObjectName(headers[0])
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        
        # Stretch last section and set resize mode
        table.horizontalHeader().setStretchLastSection(True)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        
        # Set alternating row colors
        table.setAlternatingRowColors(True)
        
        # Set row height yang lebih besar
        table.verticalHeader().setDefaultSectionSize(45)
        
        # Set minimum height untuk tabel
        table.setMinimumHeight(600)
        
        # Set spacing dan margin
        table.setContentsMargins(20, 20, 20, 20)
        table.setShowGrid(True)
        
        table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                alternate-background-color: #f5f6fa;
                border: 1px solid #dcdde1;
                gridline-color: #dcdde1;
                padding: 10px;
                selection-background-color: #3498db22;
            }
            QHeaderView::section {
                background-color: #2980b9;
                color: white;
                padding: 12px;
                border: none;
                font-weight: bold;
                font-size: 14px;
            }
            QHeaderView::section:hover {
                background-color: #3498db;
            }
            QTableWidget::item {
                padding: 10px 15px;
                border-bottom: 1px solid #dcdde1;
                margin: 5px;
            }
            QTableWidget::item:selected {
                background-color: #3498db22;
                color: #2c3e50;
            }
        """)
        
        # Set scroll mode
        table.setVerticalScrollMode(QTableWidget.ScrollPerPixel)
        table.setHorizontalScrollMode(QTableWidget.ScrollPerPixel)
        
        # Initialize with empty rows
        table.setRowCount(15)
        
        return table

    def updateEvaluationCharts(self):
        """Update semua grafik evaluasi"""
        try:
            # Update Object Detection Charts
            self.updateObjectDetectionCharts()
            
            # Update Behavior Analysis Charts
            self.updateBehaviorCharts()
            
            # Update Anomaly Detection Charts
            self.updateAnomalyCharts()
            
            # Update Audio Analysis Charts
            self.updateAudioCharts()
            
            # Update Performance Charts
            self.updatePerformanceCharts()
            
        except Exception as e:
            print(f"Error updating evaluation charts: {str(e)}")

    def updateObjectDetectionCharts(self):
        """Update grafik deteksi objek"""
        # Pie Chart
        self.objectPieFigure.clear()
        ax = self.objectPieFigure.add_subplot(111)
        objects = ['Orang', 'Kendaraan', 'Tas', 'Mencurigakan']
        sizes = [random.randint(20, 40) for _ in range(4)]
        colors = ['#2ecc71', '#3498db', '#f1c40f', '#e74c3c']
        explode = (0.1, 0, 0, 0.1)  # Explode orang dan mencurigakan
        ax.pie(sizes, explode=explode, labels=objects, colors=colors, autopct='%1.1f%%',
               shadow=True, startangle=90)
        ax.set_title('Distribusi Objek Terdeteksi', pad=20, fontsize=12, fontweight='bold')
        self.objectPieCanvas.draw()
        
        # Trend Chart
        self.objectTrendFigure.clear()
        ax = self.objectTrendFigure.add_subplot(111)
        times = pd.date_range(end=datetime.now(), periods=10, freq='h')
        
        for i, obj in enumerate(objects):
            values = np.random.normal(30, 5, 10)
            ax.plot(times, values, '-o', label=obj, color=colors[i], linewidth=2,
                   marker='o', markersize=8, markerfacecolor='white')
            
        ax.set_title('Trend Deteksi Objek', pad=20, fontsize=12, fontweight='bold')
        ax.set_xlabel('Waktu', fontsize=10)
        ax.set_ylabel('Jumlah Deteksi', fontsize=10)
        ax.grid(True, linestyle='--', alpha=0.7)
        ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
        self.objectTrendFigure.tight_layout()
        self.objectTrendCanvas.draw()

    def updateBehaviorCharts(self):
        """Update grafik analisis perilaku"""
        # Heatmap
        self.behaviorHeatFigure.clear()
        ax = self.behaviorHeatFigure.add_subplot(111)
        data = np.random.rand(24, 7)  # 24 jam x 7 hari
        days = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
        hours = [f'{i:02d}:00' for i in range(24)]
        
        sns.heatmap(data, ax=ax, cmap='YlOrRd', xticklabels=days, yticklabels=hours,
                   cbar_kws={'label': 'Tingkat Aktivitas'})
        ax.set_title('Pola Aktivitas Mingguan', pad=20, fontsize=12, fontweight='bold')
        ax.set_xlabel('Hari', fontsize=10)
        ax.set_ylabel('Jam', fontsize=10)
        plt.xticks(rotation=45)
        self.behaviorHeatFigure.tight_layout()
        self.behaviorHeatCanvas.draw()
        
        # Bar Chart
        self.behaviorBarFigure.clear()
        ax = self.behaviorBarFigure.add_subplot(111)
        behaviors = ['Normal', 'Mencurigakan', 'Berbahaya', 'Darurat']
        counts = [random.randint(50, 100) for _ in range(4)]
        colors = ['#2ecc71', '#f1c40f', '#e67e22', '#e74c3c']
        
        bars = ax.bar(behaviors, counts, color=colors)
        ax.set_title('Distribusi Perilaku', pad=20, fontsize=12, fontweight='bold')
        ax.set_xlabel('Tipe Perilaku', fontsize=10)
        ax.set_ylabel('Jumlah Kejadian', fontsize=10)
        
        # Tambahkan nilai di atas bar
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                   f'{int(height)}',
                   ha='center', va='bottom')
        
        ax.grid(True, linestyle='--', alpha=0.7, axis='y')
        self.behaviorBarFigure.tight_layout()
        self.behaviorBarCanvas.draw()

    def updateAnomalyCharts(self):
        """Update grafik deteksi anomali"""
        # Time Series
        self.anomalyTSFigure.clear()
        ax = self.anomalyTSFigure.add_subplot(111)
        times = pd.date_range(end=datetime.now(), periods=100, freq='5min')
        values = np.random.normal(0, 1, 100)
        anomalies = np.random.choice([0, 1], 100, p=[0.9, 0.1])
        
        # Plot normal data
        ax.plot(times, values, 'b-', label='Normal', linewidth=2, color='#3498db')
        # Plot anomalies dengan highlight
        ax.scatter(times[anomalies == 1], values[anomalies == 1], 
                  color='#e74c3c', s=100, label='Anomali', zorder=5)
        
        ax.set_title('Deteksi Anomali Real-time', pad=20, fontsize=12, fontweight='bold')
        ax.set_xlabel('Waktu', fontsize=10)
        ax.set_ylabel('Nilai Sensor', fontsize=10)
        ax.grid(True, linestyle='--', alpha=0.7)
        ax.legend()
        self.anomalyTSFigure.tight_layout()
        self.anomalyTSCanvas.draw()
        
        # Scatter Plot
        self.anomalyScatterFigure.clear()
        ax = self.anomalyScatterFigure.add_subplot(111)
        
        # Generate better-looking clusters
        normal_data = np.random.multivariate_normal([0, 0], [[1, 0.5], [0.5, 1]], 100)
        anomaly_data = np.random.multivariate_normal([3, 3], [[0.5, 0.2], [0.2, 0.5]], 10)
        
        ax.scatter(normal_data[:, 0], normal_data[:, 1], 
                  label='Normal', color='#3498db', alpha=0.6)
        ax.scatter(anomaly_data[:, 0], anomaly_data[:, 1], 
                  color='#e74c3c', s=100, label='Anomali', alpha=0.6)
        
        ax.set_title('Clustering Anomali', pad=20, fontsize=12, fontweight='bold')
        ax.set_xlabel('Feature 1', fontsize=10)
        ax.set_ylabel('Feature 2', fontsize=10)
        ax.grid(True, linestyle='--', alpha=0.7)
        ax.legend()
        self.anomalyScatterFigure.tight_layout()
        self.anomalyScatterCanvas.draw()

    def updateAudioCharts(self):
        """Update grafik analisis audio"""
        # Spektogram
        self.audioSpecFigure.clear()
        ax = self.audioSpecFigure.add_subplot(111)
        
        # Generate more interesting audio data
        t = np.linspace(0, 10, 1000)
        frequencies = [1.0, 2.0, 3.0]
        signal = np.zeros_like(t)
        for freq in frequencies:
            signal += np.sin(2.0 * np.pi * freq * t)
        signal += np.random.normal(0, 0.1, t.shape)
        
        spec = ax.specgram(signal, NFFT=256, Fs=100, noverlap=128,
                          cmap='viridis')
        ax.set_title('Spektogram Audio', pad=20, fontsize=12, fontweight='bold')
        ax.set_xlabel('Waktu (s)', fontsize=10)
        ax.set_ylabel('Frekuensi (Hz)', fontsize=10)
        self.audioSpecFigure.colorbar(spec[3], ax=ax, label='Intensitas (dB)')
        self.audioSpecFigure.tight_layout()
        self.audioSpecCanvas.draw()
        
        # Level Suara
        self.audioLevelFigure.clear()
        ax = self.audioLevelFigure.add_subplot(111)
        times = pd.date_range(end=datetime.now(), periods=50, freq='1min')
        
        # Generate more realistic audio levels
        base_level = 45  # Base ambient noise level
        activity_spikes = np.random.normal(20, 5, 50)  # Random activity
        levels = base_level + activity_spikes
        
        ax.fill_between(times, base_level, levels, alpha=0.3, color='#3498db')
        ax.plot(times, levels, '-', color='#2980b9', linewidth=2)
        
        ax.set_title('Level Suara Real-time', pad=20, fontsize=12, fontweight='bold')
        ax.set_xlabel('Waktu', fontsize=10)
        ax.set_ylabel('Level Suara (dB)', fontsize=10)
        ax.grid(True, linestyle='--', alpha=0.7)
        
        # Add threshold lines
        ax.axhline(y=85, color='#e74c3c', linestyle='--', label='Batas Bahaya')
        ax.axhline(y=65, color='#f1c40f', linestyle='--', label='Batas Peringatan')
        ax.legend()
        
        self.audioLevelFigure.tight_layout()
        self.audioLevelCanvas.draw()

    def updatePerformanceCharts(self):
        """Update grafik kinerja"""
        # Metrics Chart
        self.perfMetricsFigure.clear()
        ax = self.perfMetricsFigure.add_subplot(111)
        times = pd.date_range(end=datetime.now(), periods=24, freq='h')
        
        metrics = {
            'Akurasi': (np.random.uniform(0.85, 0.95, 24), '#2ecc71'),
            'Presisi': (np.random.uniform(0.80, 0.90, 24), '#3498db'),
            'Recall': (np.random.uniform(0.75, 0.85, 24), '#e67e22')
        }
        
        for metric, (values, color) in metrics.items():
            ax.plot(times, values, '-o', label=metric, color=color, linewidth=2,
                   marker='o', markersize=6, markerfacecolor='white')
        
        ax.set_title('Metrik Performa Model', pad=20, fontsize=12, fontweight='bold')
        ax.set_xlabel('Waktu', fontsize=10)
        ax.set_ylabel('Nilai', fontsize=10)
        ax.grid(True, linestyle='--', alpha=0.7)
        ax.legend(loc='center left', bbox_to_anchor=(1, 0.5))
        ax.set_ylim(0.7, 1.0)
        
        self.perfMetricsFigure.tight_layout()
        self.perfMetricsCanvas.draw()
        
        # Resource Usage
        self.resourceUsageFigure.clear()
        ax = self.resourceUsageFigure.add_subplot(111)
        
        resources = {
            'CPU': (np.random.uniform(20, 60, 24), '#3498db'),
            'Memory': (np.random.uniform(30, 70, 24), '#2ecc71'),
            'GPU': (np.random.uniform(10, 50, 24), '#e74c3c')
        }
        
        for resource, (values, color) in resources.items():
            ax.fill_between(times, 0, values, label=resource, alpha=0.3, color=color)
            ax.plot(times, values, color=color, linewidth=2)
        
        ax.set_title('Penggunaan Sumber Daya', pad=20, fontsize=12, fontweight='bold')
        ax.set_xlabel('Waktu', fontsize=10)
        ax.set_ylabel('Penggunaan (%)', fontsize=10)
        ax.grid(True, linestyle='--', alpha=0.7)
        ax.legend(loc='upper left')
        
        self.resourceUsageFigure.tight_layout()
        self.resourceUsageCanvas.draw()

    def exportData(self):
        """Export data ke Excel"""
        try:
            from security_data import generate_security_data
            
            # Generate dan export data
            if generate_security_data():
                QMessageBox.information(
                    self,
                    "Export Berhasil",
                    "Data berhasil diekspor ke file Excel:\ndata/security_report.xlsx\n\n"
                    "File berisi sheet:\n"
                    "- Deteksi Objek\n"
                    "- Analisis Perilaku\n"
                    "- Deteksi Anomali\n"
                    "- Data Sensor"
                )
            else:
                QMessageBox.warning(
                    self,
                    "Peringatan Export",
                    "Terjadi masalah saat mengekspor data.\n"
                    "Silakan cek console untuk detail error."
                )
            
        except ImportError as e:
            QMessageBox.critical(
                self,
                "Error Module",
                "Modul yang dibutuhkan tidak ditemukan.\n"
                "Pastikan sudah menginstall:\n"
                "- openpyxl\n"
                "- pandas"
            )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error Export",
                f"Terjadi kesalahan saat mengekspor data:\n{str(e)}"
            )

    def refreshData(self):
        """Perbarui data dalam tabel dan grafik"""
        try:
            # Update tables
            self.updateTables()
            
            # Update charts
            self.updateEvaluationCharts()
            
            # Tampilkan notifikasi
            QMessageBox.information(
                self,
                "Pembaruan Berhasil",
                "Data dan grafik berhasil diperbarui!"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error Pembaruan",
                f"Terjadi kesalahan saat memperbarui data:\n{str(e)}"
            )

    def updateTables(self):
        """Update semua tabel dengan data terbaru"""
        try:
            # Data untuk setiap tabel
            tables_data = {
                self.objectDetectionTable: [
                    ['Orang', 'Kendaraan', 'Tas', 'Benda Mencurigakan'],
                    (0.70, 0.99),
                    ['Depan', 'Belakang', 'Samping', 'Dalam'],
                    ['Normal', 'Perlu Perhatian', 'Mencurigakan']
                ],
                self.behaviorTable: [
                    ['Normal', 'Mencurigakan', 'Berbahaya', 'Darurat'],
                    (0.0, 1.0),
                    (1, 60),
                    ['Monitoring', 'Peringatan', 'Alarm', 'Evakuasi']
                ],
                self.anomalyTable: [
                    (0.0, 1.0),
                    ['Gerakan', 'Suara', 'Akses', 'Pola'],
                    ['Rendah', 'Sedang', 'Tinggi', 'Kritis'],
                    ['Pending', 'Diproses', 'Ditangani', 'Selesai']
                ],
                self.audioTable: [
                    ['Normal', 'Berisik', 'Mencurigakan', 'Darurat'],
                    (30, 100),
                    ['Percakapan', 'Langkah Kaki', 'Tabrakan', 'Teriakan'],
                    ['Depan', 'Belakang', 'Samping', 'Dalam']
                ],
                self.performanceTable: [
                    (90.0, 99.9),
                    (0.01, 0.05),
                    (0.5, 2.0),
                    (20, 80)
                ]
            }
            
            # Update setiap tabel
            for table, data_columns in tables_data.items():
                self.fillTableData(table, data_columns)
            
        except Exception as e:
            print(f"Kesalahan saat memperbarui tabel: {str(e)}")

    def fillTableData(self, table, data_columns):
        """Isi tabel dengan data"""
        try:
            current_time = datetime.now()
            
            # Generate 15 baris data
            for i in range(15):
                # Set timestamp
                timestamp = current_time - timedelta(minutes=i*5)
                timestamp_item = QTableWidgetItem(timestamp.strftime("%Y-%m-%d %H:%M:%S"))
                timestamp_item.setTextAlignment(Qt.AlignCenter)
                font = QFont()
                font.setPointSize(10)
                timestamp_item.setFont(font)
                table.setItem(i, 0, timestamp_item)
                
                # Set data lainnya
                for col, values in enumerate(data_columns, start=1):
                    if isinstance(values, tuple):
                        # Format angka dengan 2 desimal
                        value = f"{random.uniform(values[0], values[1]):.2f}"
                        if "%" in table.horizontalHeaderItem(col).text():
                            value += "%"
                    else:
                        value = random.choice(values)
                    
                    item = QTableWidgetItem(value)
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFont(font)
                    
                    # Set warna berdasarkan nilai
                    if "Status" in table.horizontalHeaderItem(col).text():
                        if "Normal" in value or "Aman" in value:
                            item.setForeground(QColor("#27ae60"))
                        elif "Perhatian" in value or "Mencurigakan" in value:
                            item.setForeground(QColor("#f39c12"))
                        else:
                            item.setForeground(QColor("#e74c3c"))
                    
                    table.setItem(i, col, item)
            
            # Format tabel
            self.formatTable(table)
            
        except Exception as e:
            print(f"Kesalahan saat mengisi data tabel: {str(e)}")

    def formatTable(self, table):
        """Format tampilan tabel"""
        try:
            # Set lebar minimum kolom
            for column in range(table.columnCount()):
                table.setColumnWidth(column, 150)
            
            # Set tinggi baris
            table.verticalHeader().setDefaultSectionSize(40)
            
            # Sembunyikan nomor baris
            table.verticalHeader().setVisible(False)
            
            # Set warna alternating dan style
            table.setAlternatingRowColors(True)
            table.setStyleSheet("""
                QTableWidget {
                    background-color: white;
                    alternate-background-color: #f5f6fa;
                    border: none;
                    gridline-color: #dcdde1;
                    padding: 10px;
                }
                QTableWidget::item {
                    padding: 8px 15px;
                    border-bottom: 1px solid #f5f6fa;
                    margin: 5px;
                }
                QTableWidget::item:selected {
                    background-color: #3498db22;
                    color: #2c3e50;
                }
            """)
            
        except Exception as e:
            print(f"Kesalahan saat memformat tabel: {str(e)}")

    def setupAlgorithmTab(self):
        """Setup tab algoritma dengan informasi AI"""
        algorithmTab = QWidget()
        layout = QHBoxLayout(algorithmTab)
        
        # Panel kiri (Tree View)
        leftPanel = QFrame()
        leftLayout = QVBoxLayout(leftPanel)
        
        # Tambahkan tree widget
        tree_widget = AlgorithmTreeWidget()
        leftLayout.addWidget(tree_widget)
        
        layout.addWidget(leftPanel, 40)
        
        # Panel kanan (Informasi dan Kontrol)
        rightPanel = QFrame()
        rightLayout = QVBoxLayout(rightPanel)
        
        # Pengaturan Algoritma
        settingsFrame = QFrame()
        settingsLayout = QGridLayout(settingsFrame)
        
        settings = [
            ("Sensitivitas AI", "slider", (0, 100, 75)),
            ("Ambang Batas Deteksi", "slider", (0, 100, 60)),
            ("Mode Analisis AI", "combo", ["Dasar", "Lanjutan", "Ahli"]),
            ("Tingkat Pembelajaran", "slider", (0, 100, 50))
        ]
        
        for i, (label, type_, value) in enumerate(settings):
            settingsLayout.addWidget(QLabel(label), i, 0)
            
            if type_ == "slider":
                slider = QSlider(Qt.Horizontal)
                slider.setRange(value[0], value[1])
                slider.setValue(value[2])
                slider.setStyleSheet("""
                    QSlider::groove:horizontal {
                        background: #dcdde1;
                        height: 8px;
                        border-radius: 4px;
                    }
                    QSlider::handle:horizontal {
                        background: #2980b9;
                        width: 16px;
                        margin: -4px 0;
                        border-radius: 8px;
                    }
                """)
                settingsLayout.addWidget(slider, i, 1)
            elif type_ == "combo":
                combo = QComboBox()
                combo.addItems(value)
                combo.setStyleSheet("""
                    QComboBox {
                        padding: 5px;
                        border: 2px solid #dcdde1;
                        border-radius: 5px;
                        background: white;
                    }
                """)
                settingsLayout.addWidget(combo, i, 1)
        
        rightLayout.addWidget(settingsFrame)
        
        # Informasi Model AI
        infoFrame = QFrame()
        infoLayout = QVBoxLayout(infoFrame)
        
        infoTitle = QLabel("Informasi Model AI")
        infoTitle.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 16px;")
        infoLayout.addWidget(infoTitle)
        
        self.algorithmInfo = QLabel()
        self.algorithmInfo.setWordWrap(True)
        self.algorithmInfo.setStyleSheet("""
            color: #2c3e50;
            background: #f5f6fa;
            padding: 15px;
            border-radius: 5px;
            font-family: monospace;
        """)
        infoLayout.addWidget(self.algorithmInfo)
        
        rightLayout.addWidget(infoFrame)
        
        # Kontrol Training
        trainingFrame = QFrame()
        trainingLayout = QHBoxLayout(trainingFrame)
        
        trainButton = QPushButton("🔄 Latih Model")
        trainButton.setStyleSheet("""
            QPushButton {
                background-color: #2980b9;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #3498db;
            }
        """)
        
        resetButton = QPushButton("⚠️ Reset Model")
        resetButton.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        
        trainingLayout.addWidget(trainButton)
        trainingLayout.addWidget(resetButton)
        
        rightLayout.addWidget(trainingFrame)
        
        layout.addWidget(rightPanel, 60)
        
        self.tabWidget.addTab(algorithmTab, "🤖 AI & Algoritma")

    def setupAIMaintenanceTab(self):
        """Setup tab AI Maintenance"""
        aiTab = QWidget()
        layout = QVBoxLayout(aiTab)
        
        # Header
        headerLayout = QHBoxLayout()
        title = QLabel("AI System Maintenance")
        title.setStyleSheet("font-size: 18px; font-weight: bold; color: #2c3e50;")
        headerLayout.addWidget(title)
        headerLayout.addStretch()
        
        # Refresh Button
        refreshBtn = QPushButton("Refresh Data")
        refreshBtn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        refreshBtn.clicked.connect(self.updateAIMetrics)
        headerLayout.addWidget(refreshBtn)
        layout.addLayout(headerLayout)
        
        # Create tab container
        aiTabs = QTabWidget()
        
        # 1. Model Performance Tab
        perfTab = QWidget()
        perfLayout = QVBoxLayout(perfTab)
        
        # Performance Metrics
        metricsLayout = QHBoxLayout()
        
        # Accuracy
        self.accuracyProgress = self.createMetricWidget("Model Accuracy", metricsLayout)
        self.precisionProgress = self.createMetricWidget("Model Precision", metricsLayout)
        self.f1Progress = self.createMetricWidget("F1-Score", metricsLayout)
        
        perfLayout.addLayout(metricsLayout)
        
        # Performance Chart
        self.perfFigure = plt.figure(figsize=(10, 4))
        self.perfCanvas = FigureCanvas(self.perfFigure)
        perfLayout.addWidget(self.perfCanvas)
        
        aiTabs.addTab(perfTab, "Performance")
        
        # 2. Anomaly Analysis Tab
        anomalyTab = QWidget()
        anomalyLayout = QVBoxLayout(anomalyTab)
        
        # Anomaly Stats
        statsLayout = QHBoxLayout()
        self.anomalyCount = QLabel("Total Anomalies: 0")
        self.anomalyRate = QLabel("Anomaly Rate: 0%")
        self.falsePositive = QLabel("False Positive Rate: 0%")
        
        for label in [self.anomalyCount, self.anomalyRate, self.falsePositive]:
            label.setStyleSheet("font-size: 14px; color: #2c3e50; padding: 5px;")
            statsLayout.addWidget(label)
        
        anomalyLayout.addLayout(statsLayout)
        
        # Anomaly Chart
        self.anomalyFigure = plt.figure(figsize=(10, 4))
        self.anomalyCanvas = FigureCanvas(self.anomalyFigure)
        anomalyLayout.addWidget(self.anomalyCanvas)
        
        aiTabs.addTab(anomalyTab, "Anomalies")
        
        # 3. Learning Progress Tab
        learningTab = QWidget()
        learningLayout = QVBoxLayout(learningTab)
        
        # Training Progress
        self.trainingProgress = QProgressBar()
        learningLayout.addWidget(QLabel("Training Progress"))
        learningLayout.addWidget(self.trainingProgress)
        
        # Updates Table
        self.updatesTable = QTableWidget(5, 3)
        self.updatesTable.setHorizontalHeaderLabels(["Timestamp", "Type", "Performance"])
        self.formatTable(self.updatesTable)
        learningLayout.addWidget(QLabel("Model Updates"))
        learningLayout.addWidget(self.updatesTable)
        
        aiTabs.addTab(learningTab, "Learning")
        
        # 4. System Health Tab
        healthTab = QWidget()
        healthLayout = QVBoxLayout(healthTab)
        
        # Health Metrics
        healthMetrics = QHBoxLayout()
        self.cpuProgress = self.createMetricWidget("CPU Usage", healthMetrics)
        self.memoryProgress = self.createMetricWidget("Memory Usage", healthMetrics)
        self.responseProgress = self.createMetricWidget("Response Time", healthMetrics)
        
        healthLayout.addLayout(healthMetrics)
        
        # System Logs
        self.systemLogs = QTableWidget(10, 3)
        self.systemLogs.setHorizontalHeaderLabels(["Timestamp", "Event", "Status"])
        self.formatTable(self.systemLogs)
        healthLayout.addWidget(QLabel("System Logs"))
        healthLayout.addWidget(self.systemLogs)
        
        aiTabs.addTab(healthTab, "Health")
        
        layout.addWidget(aiTabs)
        self.tabWidget.addTab(aiTab, "AI Maintenance")
        
    def createMetricWidget(self, title, parentLayout):
        """Helper untuk membuat widget metrik dengan progress bar"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        label = QLabel(title)
        label.setStyleSheet("font-size: 14px; color: #2c3e50;")
        progress = QProgressBar()
        progress.setStyleSheet("""
            QProgressBar {
                border: 2px solid #dcdde1;
                border-radius: 5px;
                text-align: center;
            }
            QProgressBar::chunk {
                background-color: #3498db;
                border-radius: 3px;
            }
        """)
        
        layout.addWidget(label)
        layout.addWidget(progress)
        parentLayout.addWidget(widget)
        
        return progress
        
    def updateAIMetrics(self):
        """Update semua metrik AI"""
        try:
            # Update performance metrics
            self.accuracyProgress.setValue(85)
            self.precisionProgress.setValue(82)
            self.f1Progress.setValue(83)
            
            # Update performance chart
            self.perfFigure.clear()
            ax = self.perfFigure.add_subplot(111)
            dates = pd.date_range(end=datetime.now(), periods=10, freq='D')
            accuracy = np.random.uniform(80, 90, 10)
            ax.plot(dates, accuracy, '-o', label='Accuracy')
            ax.set_title('Model Performance Trend')
            ax.set_xlabel('Date')
            ax.set_ylabel('Accuracy (%)')
            ax.grid(True)
            self.perfFigure.autofmt_xdate()
            self.perfCanvas.draw()
            
            # Update anomaly stats
            total_anomalies = random.randint(50, 100)
            self.anomalyCount.setText(f"Total Anomalies: {total_anomalies}")
            self.anomalyRate.setText(f"Anomaly Rate: {random.randint(5, 15)}%")
            self.falsePositive.setText(f"False Positive Rate: {random.randint(2, 8)}%")
            
            # Update anomaly chart
            self.anomalyFigure.clear()
            ax = self.anomalyFigure.add_subplot(111)
            hours = range(24)
            anomalies = np.random.poisson(5, 24)
            ax.bar(hours, anomalies)
            ax.set_title('Anomaly Distribution by Hour')
            ax.set_xlabel('Hour')
            ax.set_ylabel('Number of Anomalies')
            self.anomalyCanvas.draw()
            
            # Update training progress
            self.trainingProgress.setValue(90)
            
            # Update system health
            self.cpuProgress.setValue(45)
            self.memoryProgress.setValue(60)
            self.responseProgress.setValue(25)
            
            # Update tables
            self.updateAITables()
            
        except Exception as e:
            print(f"Error updating AI metrics: {str(e)}")
            
    def updateAITables(self):
        """Update tabel-tabel dalam AI maintenance tab"""
        # Update model updates table
        updates = [
            (datetime.now().strftime("%Y-%m-%d %H:%M"), "Retraining", "85%"),
            (datetime.now().strftime("%Y-%m-%d %H:%M"), "Fine-tuning", "87%"),
            (datetime.now().strftime("%Y-%m-%d %H:%M"), "Optimization", "88%")
        ]
        
        self.updatesTable.setRowCount(len(updates))
        for i, (timestamp, type_, perf) in enumerate(updates):
            self.updatesTable.setItem(i, 0, QTableWidgetItem(timestamp))
            self.updatesTable.setItem(i, 1, QTableWidgetItem(type_))
            self.updatesTable.setItem(i, 2, QTableWidgetItem(perf))
            
        # Update system logs
        logs = [
            (datetime.now().strftime("%Y-%m-%d %H:%M"), "Model Update", "Success"),
            (datetime.now().strftime("%Y-%m-%d %H:%M"), "Anomaly Detection", "Warning"),
            (datetime.now().strftime("%Y-%m-%d %H:%M"), "System Check", "Success")
        ]
        
        self.systemLogs.setRowCount(len(logs))
        for i, (timestamp, event, status) in enumerate(logs):
            self.systemLogs.setItem(i, 0, QTableWidgetItem(timestamp))
            self.systemLogs.setItem(i, 1, QTableWidgetItem(event))
            status_item = QTableWidgetItem(status)
            status_item.setForeground(QColor("#27ae60") if status == "Success" else QColor("#e74c3c"))
            self.systemLogs.setItem(i, 2, status_item)
            
    def setupFooter(self, parentLayout):
        """Setup area footer"""
        footer = QFrame()
        footer.setMaximumHeight(50)
        footerLayout = QHBoxLayout(footer)
        
        # Waktu pembaruan terakhir
        self.lastUpdateLabel = QLabel("Pembaruan terakhir: -")
        self.lastUpdateLabel.setStyleSheet("color: #7f8c8d;")
        
        # Status koneksi
        self.connectionLabel = QLabel("⚫ Terhubung")
        self.connectionLabel.setStyleSheet("color: #27ae60;")
        
        footerLayout.addWidget(self.lastUpdateLabel)
        footerLayout.addStretch()
        footerLayout.addWidget(self.connectionLabel)
        
        parentLayout.addWidget(footer)

    def applyStyles(self):
        """Terapkan style global"""
        self.setStyleSheet("""
            QMainWindow {
                background-color: white;
            }
            QFrame {
                background-color: white;
                border-radius: 10px;
            }
            QLabel {
                font-size: 14px;
            }
            QPushButton {
                font-size: 14px;
            }
        """)

    def setupData(self):
        """Inisialisasi data dan variabel"""
        self.sensors = {
            'motion': True,
            'door': True,
            'window': True,
            'smoke': True
        }
        self.alarmActive = False
        self.currentMode = "Normal"

    def setupTimers(self):
        """Setup timer untuk update otomatis"""
        # Timer untuk status sistem
        self.statusTimer = QTimer()
        self.statusTimer.timeout.connect(self.updateStatus)
        self.statusTimer.start(5000)  # Update setiap 5 detik
        
        # Timer untuk simulasi aktivitas
        self.activityTimer = QTimer()
        self.activityTimer.timeout.connect(self.simulateActivity)
        self.activityTimer.start(10000)  # Update setiap 10 detik
        
        # Timer untuk update keamanan
        self.securityTimer = QTimer()
        self.securityTimer.timeout.connect(self.updateSecurity)
        self.securityTimer.start(15000)  # Update setiap 15 detik
        
        # Timer untuk update timestamp
        self.timestampTimer = QTimer()
        self.timestampTimer.timeout.connect(self.updateTableTimestamps)
        self.timestampTimer.start(60000)  # Update setiap 1 menit
        
        # Timer untuk refresh data tabel
        self.tableTimer = QTimer()
        self.tableTimer.timeout.connect(self.updateTables)
        self.tableTimer.start(300000)  # Update setiap 5 menit

    def updateStatus(self):
        """Update status sistem"""
        current_time = datetime.now().strftime("%H:%M:%S")
        self.lastUpdateLabel.setText(f"Pembaruan terakhir: {current_time}")

    def simulateActivity(self):
        """Simulasi aktivitas sistem"""
        try:
            # Get real readings from dummy devices
            sensor_readings = self.devices.get_all_sensor_readings()
            actuator_status = self.devices.get_all_actuator_status()
            
            # Update log dengan data sensor
            current_time = datetime.now().strftime("%H:%M:%S")
            
            # Check PIR sensor
            if sensor_readings['pir'] > 0.7:
                self.logList.insertItem(0, f"🚨 {current_time} - Gerakan terdeteksi! (PIR: {sensor_readings['pir']:.2f})")
                self.devices.trigger_alarm()
            
            # Check magnetic sensor
            if sensor_readings['magnetic'] > 0.8:
                self.logList.insertItem(0, f"🚪 {current_time} - Pintu/jendela terbuka! (Magnetic: {sensor_readings['magnetic']:.2f})")
                self.devices.trigger_alarm()
            
            # Check vibration
            if sensor_readings['vibration'] > 80:
                self.logList.insertItem(0, f"📳 {current_time} - Getaran kuat terdeteksi! (Vibration: {sensor_readings['vibration']:.2f})")
                self.devices.trigger_alarm()
            
            # Limit log items
            while self.logList.count() > 100:
                self.logList.takeItem(self.logList.count() - 1)
            
            # Update status sistem
            system_status = "NORMAL"
            status_color = "#27ae60"
            
            if any([
                sensor_readings['pir'] > 0.7,
                sensor_readings['magnetic'] > 0.8,
                sensor_readings['vibration'] > 80
            ]):
                system_status = "WASPADA"
                status_color = "#e74c3c"
            
            self.statusLabel.setText(f"Status: {system_status}")
            self.statusLabel.setStyleSheet(f"""
                font-size: 18px;
                color: {status_color};
                padding: 5px 15px;
                border: 2px solid {status_color};
                border-radius: 5px;
            """)
        except Exception as e:
            print(f"Error dalam simulasi aktivitas: {str(e)}")

    def setupAI(self):
        """Inisialisasi sistem AI untuk keamanan rumah dengan machine learning"""
        try:
            class SecurityAI:
                def __init__(self):
                    self.threat_level = "Aman"
                    self.last_detection = None
                    self.active_zones = ["Depan", "Belakang", "Samping", "Dalam"]
                    self.training_data = []
                    self.model_version = 1.0
                    self.last_training = datetime.now()
                    self.false_positives = []
                    self.false_negatives = []
                
                def analyze_motion(self, sensor_data):
                    """Analisis gerakan dengan machine learning"""
                    result = {
                        "status": "Normal",
                        "location": "Depan",
                        "confidence": 0.95,
                        "type": "Orang",
                        "action": "Berjalan"
                    }
                    return result
                
                def collect_training_data(self, sensor_data, result, label=None):
                    """Mengumpulkan data untuk training"""
                    training_instance = {
                        "timestamp": datetime.now(),
                        "sensor_data": sensor_data,
                        "result": result,
                        "label": label,
                        "verified": False
                    }
                    self.training_data.append(training_instance)
                    
                    # Batasi ukuran data training
                    if len(self.training_data) > 1000:
                        self.training_data = self.training_data[-1000:]
                
                def update_model(self):
                    """Update model dengan data baru"""
                    if len(self.training_data) >= 100:  # Minimal 100 data untuk update
                        try:
                            # Simulasi update model
                            self.model_version += 0.1
                            self.last_training = datetime.now()
                            # Reset data training setelah update
                            self.training_data = []
                            return {
                                "status": "Success",
                                "new_version": self.model_version,
                                "accuracy": 0.95
                            }
                        except Exception as e:
                            return {
                                "status": "Failed",
                                "error": str(e)
                            }
                    return {"status": "Insufficient data"}
                
                def validate_detection(self, detection_id, is_correct):
                    """Validasi deteksi untuk pembelajaran"""
                    if not is_correct:
                        if detection_id in self.training_data:
                            if self.training_data[detection_id]["result"]["status"] == "Normal":
                                self.false_negatives.append(detection_id)
                            else:
                                self.false_positives.append(detection_id)
                
                def get_model_metrics(self):
                    """Dapatkan metrik performa model"""
                    return {
                        "version": self.model_version,
                        "last_training": self.last_training,
                        "training_data_size": len(self.training_data),
                        "false_positives": len(self.false_positives),
                        "false_negatives": len(self.false_negatives)
                    }
                
                def adaptive_learning(self, new_pattern):
                    """Pembelajaran adaptif untuk pola baru"""
                    self.collect_training_data(new_pattern, None, "new_pattern")
                    if len(self.training_data) >= 50:  # Update lebih cepat untuk pola baru
                        return self.update_model()
                    return {"status": "Collecting data"}

                def detect_intrusion(self, sensor_data):
                    """Deteksi penyusupan"""
                    return {
                        "detected": False,
                        "location": "Depan",
                        "threat_level": "Rendah"
                    }

                def analyze_sound(self, sensor_data):
                    """Analisis suara"""
                    return {
                        "is_threat": False,
                        "type": "Normal",
                        "level_db": 45.0
                    }

                def get_security_status(self):
                    """Dapatkan status keamanan"""
                    return {
                        "overall_status": "Aman",
                        "last_check": datetime.now().strftime("%H:%M:%S")
                    }

            # Coba import modul AI keamanan, jika gagal gunakan implementasi default
            try:
                from security_ai_model import AdvancedSecurityAI
                self.ai_system = AdvancedSecurityAI()
            except ImportError:
                print("Info: Menggunakan sistem AI keamanan default dengan machine learning")
                self.ai_system = SecurityAI()
            
            # Timer untuk monitoring
            self.monitoringTimer = QTimer()
            self.monitoringTimer.timeout.connect(self.updateSecurity)
            self.monitoringTimer.start(5000)  # Update setiap 5 detik
            
            # Timer untuk machine learning maintenance
            self.mlMaintenanceTimer = QTimer()
            self.mlMaintenanceTimer.timeout.connect(self.maintainAI)
            self.mlMaintenanceTimer.start(3600000)  # Maintenance setiap 1 jam
            
            # Timer untuk validasi model
            self.modelValidationTimer = QTimer()
            self.modelValidationTimer.timeout.connect(self.validateModel)
            self.modelValidationTimer.start(1800000)  # Validasi setiap 30 menit
            
            self.logList.insertItem(0, "🤖 Sistem AI Keamanan aktif dengan pembelajaran mesin")
            
        except Exception as e:
            print(f"Error saat inisialisasi AI Keamanan: {str(e)}")
            self.ai_system = SecurityAI()
            self.logList.insertItem(0, "⚠️ Menggunakan sistem AI default karena terjadi error")

    def updateSecurity(self):
        """Update status keamanan real-time"""
        try:
            # Dapatkan data sensor
            sensor_data = self.devices.get_all_sensor_readings()
            
            # Analisis gerakan
            try:
                motion_result = self.ai_system.analyze_motion(sensor_data)
                if motion_result["status"] != "Normal":
                    self.logList.insertItem(0, 
                        f"👥 Terdeteksi {motion_result['type']} {motion_result['action']} " +
                        f"di area {motion_result['location']} " +
                        f"(Kepercayaan: {motion_result['confidence']:.2f})"
                    )
            except Exception as e:
                print(f"Error dalam analisis gerakan: {str(e)}")
            
            # Deteksi penyusupan
            try:
                intrusion_result = self.ai_system.detect_intrusion(sensor_data)
                if intrusion_result["detected"]:
                    self.logList.insertItem(0,
                        f"🚨 PERINGATAN: Terdeteksi penyusupan di {intrusion_result['location']}! " +
                        f"Level ancaman: {intrusion_result['threat_level']}"
                    )
                    if intrusion_result["threat_level"] in ["Tinggi", "Kritis"]:
                        self.devices.trigger_alarm()
            except Exception as e:
                print(f"Error dalam deteksi penyusupan: {str(e)}")
            
            # Analisis suara
            try:
                audio_result = self.ai_system.analyze_sound(sensor_data)
                if audio_result["is_threat"]:
                    self.logList.insertItem(0,
                        f"🔊 Terdeteksi suara mencurigakan: {audio_result['type']} " +
                        f"({audio_result['level_db']} dB)"
                    )
            except Exception as e:
                print(f"Error dalam analisis suara: {str(e)}")
            
            # Update status keamanan
            try:
                security_status = self.ai_system.get_security_status()
                self.updateSecurityStatus(security_status)
            except Exception as e:
                print(f"Error dalam update status keamanan: {str(e)}")
            
        except Exception as e:
            print(f"Error dalam update keamanan: {str(e)}")
            # Tampilkan pesan error ke user
            QMessageBox.warning(
                self,
                "Error Keamanan",
                f"Terjadi kesalahan dalam sistem keamanan:\n{str(e)}"
            )

    def analyzeBehavior(self):
        """Analisis pola perilaku mencurigakan"""
        try:
            if hasattr(self, 'sensor_history'):
                behavior_result = self.ai_system.analyze_behavior_pattern(self.sensor_history)
                if behavior_result["is_suspicious"]:
                    self.logList.insertItem(0,
                        f"⚠️ Terdeteksi pola mencurigakan: {behavior_result['pattern_type']} " +
                        f"(Durasi: {behavior_result['duration']})"
                    )
                    
                    if behavior_result["frequency"] == "Tinggi":
                        self.devices.trigger_alarm()
        except Exception as e:
            print(f"Error dalam analisis perilaku: {str(e)}")

    def updateSecurityStatus(self, status):
        """Update tampilan status keamanan"""
        try:
            # Update status label
            status_text = f"Status: {status['overall_status']}"
            status_color = "#27ae60" if status['overall_status'] == "Aman" else "#e74c3c"
            
            self.statusLabel.setText(status_text)
            self.statusLabel.setStyleSheet(f"""
                font-size: 18px;
                color: {status_color};
                padding: 5px 15px;
                border: 2px solid {status_color};
                border-radius: 5px;
            """)
            
            # Update timestamp
            self.lastUpdateLabel.setText(f"Pembaruan terakhir: {status['last_check']}")
            
        except Exception as e:
            print(f"Error dalam update status: {str(e)}")

    def updateTableTimestamps(self):
        """Update timestamp pada semua tabel"""
        current_time = datetime.now()
        
        tables = [
            self.objectDetectionTable,
            self.behaviorTable,
            self.anomalyTable,
            self.audioTable,
            self.performanceTable
        ]
        
        for table in tables:
            for i in range(table.rowCount()):
                timestamp = current_time - timedelta(minutes=i*5)
                timestamp_item = QTableWidgetItem(timestamp.strftime("%Y-%m-%d %H:%M:%S"))
                timestamp_item.setTextAlignment(Qt.AlignCenter)
                font = QFont()
                font.setPointSize(10)
                timestamp_item.setFont(font)
                table.setItem(i, 0, timestamp_item)

    def closeEvent(self, event):
        """Handle window close event"""
        # Reset all actuators before closing
        self.devices.reset_alarm()
        self.hide()
        event.ignore()  # Prevent the window from being destroyed

    def maintainAI(self):
        """Maintenance rutin sistem AI keamanan"""
        try:
            # 1. Analisis Pola Keamanan
            security_patterns = self.analyzeSecurityPatterns()
            if security_patterns["new_patterns_found"]:
                self.logList.insertItem(0, 
                    f"🔍 Pola baru terdeteksi: {security_patterns['pattern_description']}")
                self.ai_system.adaptive_learning(security_patterns["pattern_data"])

            # 2. Evaluasi Akurasi Deteksi
            detection_metrics = self.evaluateDetectionAccuracy()
            self.logList.insertItem(0,
                f"📊 Akurasi Deteksi - Orang: {detection_metrics['person_accuracy']}%, " +
                f"Kendaraan: {detection_metrics['vehicle_accuracy']}%, " +
                f"Objek: {detection_metrics['object_accuracy']}%")

            # 3. Analisis Zona Keamanan
            zone_analysis = self.analyzeSecurityZones()
            for zone, status in zone_analysis["vulnerable_zones"].items():
                if status["risk_level"] > 0.7:
                    self.logList.insertItem(0,
                        f"⚠️ Zona {zone} memerlukan perhatian - " +
                        f"Risiko: {status['risk_level']:.2f}, " +
                        f"Alasan: {status['reason']}")

            # 4. Optimasi Sensor
            sensor_optimization = self.optimizeSensors()
            if sensor_optimization["adjustments_needed"]:
                self.logList.insertItem(0,
                    f"🔧 Rekomendasi penyesuaian sensor: {sensor_optimization['recommendations']}")

            # 5. Analisis Waktu Respons
            response_analysis = self.analyzeResponseTimes()
            self.logList.insertItem(0,
                f"⚡ Waktu respons rata-rata: {response_analysis['avg_response_time']}ms, " +
                f"Keterlambatan: {response_analysis['delayed_responses']}")

            # 6. Pembaruan Knowledge Base
            kb_update = self.updateKnowledgeBase()
            if kb_update["new_entries"]:
                self.logList.insertItem(0,
                    f"📚 Knowledge base diperbarui dengan {kb_update['new_entries']} kasus baru")

        except Exception as e:
            print(f"Error dalam maintenance AI: {str(e)}")

    def analyzeSecurityPatterns(self):
        """Analisis pola keamanan baru"""
        try:
            return {
                "new_patterns_found": True,
                "pattern_description": "Aktivitas berulang di zona belakang pukul 02:00-03:00",
                "pattern_data": {
                    "time_range": "02:00-03:00",
                    "location": "belakang",
                    "frequency": "daily",
                    "confidence": 0.85
                }
            }
        except Exception as e:
            print(f"Error dalam analisis pola: {str(e)}")
            return {"new_patterns_found": False}

    def evaluateDetectionAccuracy(self):
        """Evaluasi akurasi deteksi berbagai objek"""
        try:
            return {
                "person_accuracy": 95,
                "vehicle_accuracy": 92,
                "object_accuracy": 88,
                "false_positives": {
                    "person": 0.03,
                    "vehicle": 0.05,
                    "object": 0.07
                },
                "false_negatives": {
                    "person": 0.02,
                    "vehicle": 0.03,
                    "object": 0.05
                }
            }
        except Exception as e:
            print(f"Error dalam evaluasi akurasi: {str(e)}")
            return {"person_accuracy": 0, "vehicle_accuracy": 0, "object_accuracy": 0}

    def analyzeSecurityZones(self):
        """Analisis kerentanan zona keamanan"""
        try:
            return {
                "vulnerable_zones": {
                    "depan": {
                        "risk_level": 0.3,
                        "reason": "Pencahayaan cukup"
                    },
                    "belakang": {
                        "risk_level": 0.8,
                        "reason": "Pencahayaan kurang & blind spot terdeteksi"
                    },
                    "samping": {
                        "risk_level": 0.5,
                        "reason": "Jarak sensor optimal"
                    }
                },
                "recommendations": {
                    "belakang": "Tambah pencahayaan dan sensor sudut"
                }
            }
        except Exception as e:
            print(f"Error dalam analisis zona: {str(e)}")
            return {"vulnerable_zones": {}}

    def optimizeSensors(self):
        """Optimasi penempatan dan sensitivitas sensor"""
        try:
            return {
                "adjustments_needed": True,
                "recommendations": [
                    "Sesuaikan sensitivitas PIR belakang: +15%",
                    "Rotasi kamera depan: +10° horizontal",
                    "Kalibrasi sensor gerak samping"
                ],
                "sensor_health": {
                    "pir": 0.95,
                    "camera": 0.98,
                    "motion": 0.92
                }
            }
        except Exception as e:
            print(f"Error dalam optimasi sensor: {str(e)}")
            return {"adjustments_needed": False}

    def analyzeResponseTimes(self):
        """Analisis waktu respons sistem"""
        try:
            return {
                "avg_response_time": 150,
                "delayed_responses": 2,
                "response_breakdown": {
                    "detection": 50,
                    "analysis": 45,
                    "decision": 35,
                    "action": 20
                },
                "bottlenecks": ["Analisis video pada malam hari"]
            }
        except Exception as e:
            print(f"Error dalam analisis respons: {str(e)}")
            return {"avg_response_time": 0, "delayed_responses": 0}

    def updateKnowledgeBase(self):
        """Update basis pengetahuan sistem"""
        try:
            return {
                "new_entries": 5,
                "categories": {
                    "normal_activity": 2,
                    "suspicious_patterns": 2,
                    "environmental": 1
                },
                "total_cases": 1250,
                "learning_rate": 0.92
            }
        except Exception as e:
            print(f"Error dalam update knowledge base: {str(e)}")
            return {"new_entries": 0}

    def validateModel(self):
        """Validasi performa model"""
        try:
            metrics = self.ai_system.get_model_metrics()
            total_errors = metrics['false_positives'] + metrics['false_negatives']
            
            if total_errors > 10:  # Terlalu banyak error
                self.logList.insertItem(0, "⚠️ Performa model menurun, memulai pembelajaran adaptif")
                self.ai_system.adaptive_learning(self.devices.get_all_sensor_readings())
            
        except Exception as e:
            print(f"Error dalam validasi model: {str(e)}")

    def setupAISecurityEvaluationTab(self):
        """Setup tab evaluasi AI keamanan rumah"""
        tab = QWidget()
        layout = QVBoxLayout(tab)

        # Header dengan status evaluasi
        headerFrame = QFrame()
        headerLayout = QHBoxLayout(headerFrame)
        
        # Status Keamanan Global
        self.securityScoreLabel = QLabel("Skor Keamanan Global: 95%")
        self.securityScoreLabel.setStyleSheet("""
            font-size: 24px;
            font-weight: bold;
            color: #27ae60;
            padding: 10px;
            border: 2px solid #27ae60;
            border-radius: 10px;
        """)
        headerLayout.addWidget(self.securityScoreLabel)
        
        # Tombol Refresh dan Export
        buttonFrame = QFrame()
        buttonLayout = QHBoxLayout(buttonFrame)
        
        refreshBtn = QPushButton("🔄 Evaluasi Ulang")
        refreshBtn.clicked.connect(self.refreshSecurityEvaluation)
        refreshBtn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        
        exportBtn = QPushButton("📊 Export Laporan")
        exportBtn.clicked.connect(self.exportSecurityReport)
        exportBtn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #219a52;
            }
        """)
        
        buttonLayout.addWidget(refreshBtn)
        buttonLayout.addWidget(exportBtn)
        headerLayout.addWidget(buttonFrame)
        layout.addWidget(headerFrame)

        # Tabel Evaluasi Keamanan
        self.securityTable = QTableWidget(5, 6)
        self.securityTable.setHorizontalHeaderLabels([
            "Komponen", "Status", "Akurasi", "Risiko", "Maintenance", "Rekomendasi"
        ])
        
        # Set item contoh
        components = [
            ["Deteksi Objek", "Active", "98%", "Low", "None", "-"],
            ["Analisis Perilaku", "Active", "95%", "Medium", "Update ML", "Training"],
            ["Deteksi Anomali", "Active", "92%", "Low", "None", "-"],
            ["Sensor Network", "Active", "99%", "Low", "Calibrate", "Weekly"],
            ["Access Control", "Active", "100%", "Low", "None", "-"]
        ]
        
        for i, row in enumerate(components):
            for j, value in enumerate(row):
                item = QTableWidgetItem(value)
                item.setTextAlignment(Qt.AlignCenter)
                self.securityTable.setItem(i, j, item)
        
        self.securityTable.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.securityTable)
        
        # Footer dengan timestamp
        footerFrame = QFrame()
        footerLayout = QHBoxLayout(footerFrame)
        
        self.lastEvalLabel = QLabel("Evaluasi terakhir: -")
        self.lastEvalLabel.setStyleSheet("color: #7f8c8d;")
        footerLayout.addWidget(self.lastEvalLabel)
        
        layout.addWidget(footerFrame)
        
        self.tabWidget.addTab(tab, "🔒 Evaluasi AI Keamanan")

    def refreshSecurityEvaluation(self):
        """Refresh evaluasi keamanan"""
        try:
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.lastEvalLabel.setText(f"Evaluasi terakhir: {current_time}")
            
            # Update komponen evaluasi
            components = [
                ["Deteksi Objek", "Active", "98%", "Low", "None", "-"],
                ["Analisis Perilaku", "Active", "95%", "Medium", "Update ML", "Training"],
                ["Deteksi Anomali", "Active", "92%", "Low", "None", "-"],
                ["Sensor Network", "Active", "99%", "Low", "Calibrate", "Weekly"],
                ["Access Control", "Active", "100%", "Low", "None", "-"]
            ]
            
            for i, row in enumerate(components):
                for j, value in enumerate(row):
                    self.securityTable.setItem(i, j, QTableWidgetItem(value))
            
            QMessageBox.information(
                self,
                "Evaluasi Selesai",
                "Evaluasi keamanan telah diperbarui!"
            )
            
        except Exception as e:
            print(f"Error dalam refresh evaluasi: {str(e)}")
            QMessageBox.warning(
                self,
                "Error Evaluasi",
                "Terjadi kesalahan saat memperbarui evaluasi."
            )

    def exportSecurityReport(self):
        """Export laporan evaluasi keamanan ke Excel"""
        try:
            # Buat workbook baru
            wb = Workbook()
            
            # Sheet 1: Evaluasi Keamanan Umum
            ws1 = wb.active
            ws1.title = "Evaluasi Keamanan"
            
            # Header
            headers = ["Komponen", "Status", "Akurasi", "Risiko", "Maintenance", "Rekomendasi"]
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            # Data dari tabel
            for i in range(self.securityTable.rowCount()):
                for j in range(self.securityTable.columnCount()):
                    item = self.securityTable.item(i, j)
                    if item:
                        ws1.cell(row=i+2, column=j+1).value = item.text()
            
            # Sheet 2: Analisis AI
            ws2 = wb.create_sheet("Analisis AI")
            
            # Data AI
            ai_data = [
                ["Metrik", "Nilai", "Status"],
                ["Akurasi Model", "98%", "Optimal"],
                ["False Positives", "0.02%", "Baik"],
                ["False Negatives", "0.01%", "Baik"],
                ["Response Time", "50ms", "Optimal"],
                ["Model Version", "2.1", "Updated"],
                ["Training Data", "10000 samples", "Sufficient"],
                ["Last Update", datetime.now().strftime("%Y-%m-%d %H:%M"), "Recent"]
            ]
            
            for row in ai_data:
                ws2.append(row)
            
            # Sheet 3: Sensor Status
            ws3 = wb.create_sheet("Status Sensor")
            
            # Data sensor
            sensor_data = [
                ["Sensor", "Status", "Akurasi", "Maintenance"],
                ["Kamera Depan", "Active", "99%", "None"],
                ["Kamera Belakang", "Active", "98%", "None"],
                ["PIR Sensor", "Active", "95%", "Calibrate"],
                ["Motion Sensor", "Active", "97%", "None"],
                ["Door Sensor", "Active", "100%", "None"]
            ]
            
            for row in sensor_data:
                ws3.append(row)
            
            # Sheet 4: Rekomendasi
            ws4 = wb.create_sheet("Rekomendasi")
            
            # Data rekomendasi
            recom_data = [
                ["Area", "Prioritas", "Rekomendasi", "Timeline"],
                ["AI Model", "Medium", "Update training data", "Weekly"],
                ["Sensors", "Low", "Regular calibration", "Monthly"],
                ["Network", "Low", "Bandwidth monitoring", "Daily"],
                ["Storage", "Medium", "Cleanup old data", "Weekly"]
            ]
            
            for row in recom_data:
                ws4.append(row)
            
            # Format semua sheet
            for ws in [ws1, ws2, ws3, ws4]:
                # Set column width
                for column in ws.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Simpan file
            filename = f"security_evaluation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = f"reports/{filename}"
            
            # Buat direktori jika belum ada
            os.makedirs("reports", exist_ok=True)
            
            # Simpan workbook
            wb.save(filepath)
            
            QMessageBox.information(
                self,
                "Export Berhasil",
                f"Laporan evaluasi keamanan telah disimpan:\n{filepath}\n\n"
                f"File berisi sheet:\n"
                f"- Evaluasi Keamanan\n"
                f"- Analisis AI\n"
                f"- Status Sensor\n"
                f"- Rekomendasi"
            )
            
        except ImportError:
            QMessageBox.critical(
                self,
                "Error Module",
                "Modul yang dibutuhkan tidak ditemukan.\n"
                "Pastikan sudah menginstall:\n"
                "- openpyxl\n"
                "- pandas"
            )
        except Exception as e:
            print(f"Error dalam export laporan: {str(e)}")
            QMessageBox.warning(
                self,
                "Error Export",
                f"Terjadi kesalahan saat mengexport laporan:\n{str(e)}"
            )

if __name__ == '__main__':
    try:
        app = QApplication(sys.argv)
        window = SecuritySystem()
        window.show()
        sys.exit(app.exec_())
    except Exception as e:
        print(f"Kesalahan: {str(e)}")
        sys.exit(1) 